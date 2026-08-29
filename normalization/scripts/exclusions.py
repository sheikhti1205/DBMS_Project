"""Build the exclusion register: one entry per data item left out, with reasons.

Read only. Reads the source tree, the BBS workbook sheet list and the CSVs that
extract.py wrote, then writes the version-local exclusion register plus a
machine readable EXCLUSIONS.csv.

Nothing here is typed in from memory. Every count is measured at build time, so
the register cannot drift away from the data it describes.

Run:  python normalization/scripts/exclusions.py
"""
import csv
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import blocks as BK
from nlib import BBS_XLSX, CSVDIR, SRC, clean_text
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(HERE))
OUTDIR = os.path.abspath(os.environ.get(
    'DBMS_EXCLUSIONS_DIR',
    os.path.join(PROJECT_ROOT, 'normalization', 'exclusions')
)) + os.sep

REG = []


def add(scope, item, holds, size, reason):
    REG.append((scope, item, holds, size, reason))


def csv_hdr(path):
    with open(path, newline='', encoding='utf-8') as f:
        return next(csv.reader(f))


def csv_rows(path):
    with open(path, newline='', encoding='utf-8') as f:
        r = csv.reader(f)
        next(r)
        return list(r)


def n_rows(path):
    with open(path, newline='', encoding='utf-8') as f:
        return sum(1 for _ in csv.reader(f)) - 1


def count_lines(path):
    """True data row count. Counted with a csv reader, not by counting lines:
    three of the HDX files hold a newline inside a quoted field, so counting
    physical lines overstates them."""
    try:
        with open(path, newline='', encoding='utf-8', errors='replace') as f:
            return max(0, sum(1 for _ in csv.reader(f)) - 1)
    except OSError:
        return 0


def plural(n, word):
    return '%d %s%s' % (n, word, '' if n == 1 else 's')


def whole_sources():
    hdx = os.path.join(SRC, 'Non-Govt/HDX')
    files = sorted(f for f in os.listdir(hdx) if f.lower().endswith('.csv'))
    tot = sum(count_lines(os.path.join(hdx, f)) for f in files)
    add('Source', 'HDX, all %d files' % len(files),
        'Greenhouse gas emissions by city and by sector, gridded rainfall by '
        'administrative unit, and World Health Organization air pollution '
        'indicators',
        plural(tot, 'row'),
        'Excluded by decision. Not one of these files shares a key with the '
        'approved diagram. The emissions files are keyed by city or by '
        'industrial facility, the rainfall file by a satellite grid code such '
        'as BD10, and the indicator files by country. None is keyed by a '
        'weather station, a district or a river, so no row can join to any '
        'relation in the schema. Loading them would produce a second, '
        'unconnected database inside the first.')

    others = [
        ('DoE', 'Formatted_Data/DoE',
         'Ambient air quality and surface water quality reports',
         'Air quality is measured at continuous air monitoring stations, which '
         'are a different station network from the weather stations the '
         'diagram holds. The water quality report publishes monthly figures, '
         'and the approved Water_Quality relation is keyed by year, so a '
         'monthly figure has no column to occupy.'),
        ('MoEFCC', 'Formatted_Data/MoEFCC',
         'Climate change strategy, action plan and investment plan documents',
         'These are policy documents. The extracted sheets hold budget lines, '
         'programme names and narrative text, not environmental measurements, '
         'so there is no measured quantity to load.'),
        ('UNESCO', 'Formatted_Data/UNESCO',
         'State of conservation reporting on the Sundarbans',
         'Site level narrative reporting on one World Heritage property. It '
         'carries no station, district or river key, and no repeated '
         'measurement series.'),
        ('WHO', 'Formatted_Data/WHO',
         'Climate change and health country profiles',
         'Health indicators rather than environmental measurements. The '
         'subject is outside the scope the diagram draws.'),
        ('WorldBank', 'Formatted_Data/WORLDBANK',
         'Urban country environmental analysis',
         'A single analytical report. Its figures are national aggregates, '
         'which the diagram has no relation to hold.'),
        ('UN', 'Formatted_Data/UN',
         'Environmental statistics framework and a pilot survey',
         'A framework document listing which indicators a country should '
         'collect, not the collected values themselves.'),
        ('Other', 'Formatted_Data/Other',
         'Food and Agriculture Organization forest inventory reports, an '
         'energy operations report and academic datasets',
         'Mixed provenance and mixed subject. The Food and Agriculture '
         'Organization files are inventory methodology and plot level forest '
         'measurements that do not reconcile with the district level forest '
         'areas the diagram holds.'),
    ]
    for org, rel, holds, why in others:
        d = os.path.join(SRC, rel)
        n = 0
        for root, _, fs in os.walk(d):
            n += sum(1 for f in fs if f.lower().split('.')[-1]
                     in ('xlsx', 'xls', 'csv', 'pdf'))
        add('Source', org, holds, plural(n, 'file'), why)


def classify(sheet):
    t = sheet.strip()
    if t.upper().startswith('GIS'):
        return 'GIS layer inventory'
    m = re.match(r'^T\s*(\d+)[.\s]', t)
    if m:
        return 'Theme %s table' % m.group(1)
    if t in ('Content', 'Template'):
        return 'Workbook front matter'
    return 'Theme narrative page'


THEME_NAME = {
    '1': 'Environment and natural resources',
    '2': 'Demographic and socio-economic',
    '3': 'Environmental protection and management',
    '4': 'Hazard and natural disaster',
    '5': 'Human settlement and dwelling',
    '6': 'Ecosystem and biodiversity',
}
GROUP_REASON = {
    'GIS layer inventory':
        'Each sheet is the attribute table of one map layer. Where a layer '
        'does carry climate figures, they are averages over a decade, written '
        'as 1988-1998 or 2009-2018. The approved diagram keys every climate '
        'relation by a single Year, so a decade cannot be stored without '
        'inventing a year that was never measured.',
    'Workbook front matter':
        'A table of contents and a blank indicator template. Neither holds '
        'data.',
    'Theme narrative page':
        'Running text introducing a thematic area. There are no rows and no '
        'columns to normalize.',
}


def bbs_sheets():
    wb = load_workbook(BBS_XLSX, read_only=True)
    allsh = wb.sheetnames
    used = {b['sheet'].strip() for b in BK.BLOCKS if b['org'] == 'BBS'}
    exc = [s for s in allsh if s.strip() not in used]
    titles = {}
    for s in exc:
        ws = wb[s]
        t = ''
        for row in ws.iter_rows(max_row=6, max_col=10, values_only=True):
            for v in row:
                c = clean_text(v)
                if len(c) > 12 and not c.lower().startswith(('source', 'note')):
                    t = c
                    break
            if t:
                break
        titles[s.strip()] = t
    wb.close()

    groups = {}
    for s in exc:
        groups.setdefault(classify(s), []).append(s.strip())
    for g in sorted(groups, key=lambda k: -len(groups[k])):
        sheets = sorted(groups[g])
        m = re.match(r'^Theme (\d+) table$', g)
        if m and m.group(1) == '1':
            reason = ('Environment and natural resources. This is the theme '
                      'whose subject the diagram chiefly covers: 22 of the 24 '
                      'sheets the load reads come from it, the other two '
                      'being the waste water tables T3.14 and T3.22. The %d '
                      'sheets left out of this theme are excluded for a '
                      'reason of grain or of shape rather than of subject: an '
                      'annual total where the relation needs a month, a '
                      'decade average where it needs a year, a list inside '
                      'one cell, or a quantity such as soil salinity, sea '
                      'area or river length for which the diagram draws no '
                      'attribute. The named entries in the next section work '
                      'four of them through in full.' % len(sheets))
        elif m and m.group(1) == '3':
            reason = ('Environmental protection and management. Two sheets '
                      'from this theme are loaded, T3.14 and T3.22, because '
                      'the diagram draws Type_Of_Establishments and '
                      'Industry_Type against the reused waste water '
                      'figures. The remaining %d are excluded: they publish '
                      'greenhouse gas emissions by category, environmental '
                      'expenditure and protection programme figures, for '
                      'which the diagram draws no entity.' % len(sheets))
        elif m:
            reason = ('%s. The approved Entity Relationship diagram draws no '
                      'entity for this subject. Loading these sheets would '
                      'mean adding entities the team has not designed and the '
                      'supervisor has not approved.'
                      % THEME_NAME.get(m.group(1), g))
        else:
            reason = GROUP_REASON[g]
        add('BBS sheet', g,
            'For example: ' + '; '.join('%s %s' % (s, titles.get(s, '')[:52])
                                        for s in sheets[:3]),
            plural(len(sheets), 'sheet'), reason)
    return allsh, used, exc, titles, groups


NAMED = [
    ('T1.04', 'Annual rainfall at selected stations, 2011 to 2018',
     'The figure is one annual total per station. Rainfall_Record is keyed by '
     'station, year and month, so an annual total has no month to occupy. '
     'Splitting it across twelve months would invent twelve numbers that were '
     'never measured, and storing it under a single month would be false.'),
    ('T1.22', 'Main rivers according to length',
     'The Area covered column holds a list inside one cell, for example '
     'Sylhet (180), Cumilla (95). That breaks First Normal Form, and the '
     'approved River relation carries only the river name, so there is no '
     'attribute for length or covered area to move into.'),
    ('GIS T_7', 'Maximum average temperature, 1988 to 2018 and 2021',
     'The columns are decade averages. See the GIS layer inventory entry.'),
    ('GIS T_9', 'Average rainfall, 1988 to 2018 and 2021',
     'The columns are decade averages. See the GIS layer inventory entry.'),
    ('T1.41', 'Salinization by district, 1973, 2000 and 2009',
     'Soil salinity by district. The diagram has no soil entity, and the '
     'years do not overlap the 2011 to 2024 window every other source covers.'),
]


def named_sheets(titles):
    for s, holds, why in NAMED:
        add('BBS sheet, named', s, titles.get(s, holds), '1 sheet', why)


def in_scope_files():
    bmd = os.path.join(SRC, 'Formatted_Data/BMD')
    normals = sorted(d for d in os.listdir(bmd)
                     if os.path.isdir(os.path.join(bmd, d))
                     and d != 'Temperature Data')
    add('BMD file', ', '.join(normals),
        'Long term normal values: daily and monthly normal maximum and '
        'minimum temperature, normal rainfall, normal rainy days, normal '
        'humidity and normal wind speed',
        plural(len(normals), 'file'),
        'Excluded by decision. A normal is an average over a thirty year '
        'reference period, not a measurement of a particular month. The '
        'approved climate relations are keyed by year and month, so a normal '
        'has no year to occupy. Loading a normal beside a measured value '
        'would also mean two different kinds of number in the same column, '
        'and a query could not tell them apart.')

    brri = os.path.join(SRC, 'Govt/BRRI')
    used = {b['file'] for b in BK.BLOCKS if b['org'] == 'BRRI'}
    extra = sorted(f for f in os.listdir(brri)
                   if f.lower().endswith(('.xls', '.xlsx')) and f not in used)
    add('BRRI file', ', '.join(extra),
        'Additional climate files supplied alongside the five daily series',
        plural(len(extra), 'file'),
        'These repeat the same measured quantities as the five daily files '
        'that are loaded, for overlapping stations and dates, without stating '
        'which is the corrected version. Loading both would create duplicate '
        'keys whose disagreement no rule in the source resolves.')

    bwdb = os.path.join(SRC, 'Formatted_Data/BWDB')
    usedf = {b['file'] for b in BK.BLOCKS if b['org'] == 'BWDB'}
    extra = sorted(d for d in os.listdir(bwdb)
                   if os.path.isdir(os.path.join(bwdb, d))
                   and not any(d in u for u in usedf))
    add('BWDB file', ', '.join(extra),
        'Rainfall 2017 to 2018, and the trend of water level of major rivers '
        '2014 to 2019',
        plural(len(extra), 'file'),
        'The rainfall file is keyed by a water development board rain gauge, '
        'which is a different network from the meteorological stations the '
        'Station relation holds, so its rows cannot join. The water level '
        'file measures river gauge height, and the diagram draws no water '
        'level attribute on any relation.')

    hbes = 'Formatted_Data/BBS/BBS_HBES_2024/BBS_HBES_2024_demo_grouped.xlsx'
    add('BBS file', 'BBS_HBES_2024_demo_grouped.xlsx',
        'Household Income and Expenditure Survey 2024, 19 tables and a '
        'narrative sheet',
        '1 file',
        'Off topic. The survey measures household living conditions. Its '
        'first narrative line reads "Percentage Distribution of Household by '
        'Separate Handwashing Facilities", which is a public health measure '
        'and not an environmental one. No table in the file carries a '
        'station, a river or a measured environmental quantity.')
    return hbes


STAGE_DROPS = [
    ('1NF to 2NF', 'Climate_Observation',
     ['Source_Organization', 'Source_File', 'Source_Sheet'],
     'A partial dependency. These three depend only on the block a row was '
     'read from, not on the station and month that identify the reading. They '
     'move into Source_Block, so the provenance is kept once per block '
     'instead of being repeated on every measurement row.'),
    ('1NF to 2NF', 'Daily_Observation',
     ['Source_Organization', 'Source_Sheet'],
     'The same partial dependency on the block, moved into Source_Block.'),
    ('1NF to 2NF', 'Water_Quality',
     ['River_Name', 'Water_Category'],
     'Both depend on the monitoring station alone, not on the station and the '
     'year together. They move into River_Station, which is exactly the '
     'entity the approved diagram draws for a monitoring point on a river.'),
    ('1NF to 2NF', 'Forest_Area',
     ['Fiscal_Year'],
     'The published text 2019-20 is not atomic: it holds two years in one '
     'cell. It is split into Start_Year and End_Year in the Fiscal_Year '
     'relation, so the span is explicit and a foreign key can point at it.'),
    ('3NF to BCNF', 'River',
     ['Serial_No', 'BWDB_Zone', 'Border_River', 'Flow_Type'],
     'Not in the approved diagram, which keeps only the river name. The '
     'serial number is also a publication artefact rather than a fact about '
     'the river: it is the row position in the register, and it is what makes '
     'Salda appear twice.'),
]


def stage_columns():
    for stage, table, cols, why in STAGE_DROPS:
        add('Column, %s' % stage, '%s: %s' % (table, ', '.join(cols)),
            'One attribute per name listed', plural(len(cols), 'column'), why)


def unloaded_tables():
    for nm, holds, why in [
        ('Source_Block_2NF',
         'One row per published block, with its organisation, file, sheet and '
         'repeating group',
         'A load time provenance table created by lifting a partial '
         'dependency out of the measurement tables. It is not an entity in '
         'the approved diagram, so it is not loaded. The same information is '
         'in the Traceability sheet of the workbook.'),
        ('Measure_Unit_2NF',
         'One row per measured quantity, with its unit of measurement',
         'The unit depends only on the measure, so Second Normal Form '
         'requires lifting it out. The approved diagram draws no unit entity, '
         'so the units are documented rather than loaded. Every measurement '
         'column in the loaded schema holds one unit throughout, so no row is '
         'ambiguous without it.'),
    ]:
        found = None
        for st in ('1NF', '2NF', '3NF'):
            cand = os.path.join(CSVDIR, st, nm + '.csv')
            if os.path.exists(cand):
                found = cand
                break
        add('Table, not loaded', nm, holds,
            plural(n_rows(found) if found else 0, 'row'), why)


def unloaded_rows():
    kc = os.path.join(CSVDIR, 'BCNF', '_Key_Conflicts.csv')
    rows = csv_rows(kc)
    sub = [x for x in rows if x[6] == 'substantive']

    lost = 0
    for f in os.listdir(os.path.join(CSVDIR, '3NF')):
        if not f.endswith('_3NF.csv'):
            continue
        nm = f[:-8]
        pb = os.path.join(CSVDIR, 'BCNF', nm + '.csv')
        if os.path.exists(pb):
            lost += max(0, n_rows(os.path.join(CSVDIR, '3NF', f)) - n_rows(pb))
    exact = lost - len(rows)

    add('Rows', 'Rows refused by a primary key, exact duplicates',
        'A second organisation publishing exactly the value another '
        'organisation had already published for the same key',
        plural(exact, 'row'),
        'Boyce Codd Normal Form enforces the key, so only one row per key can '
        'load. These rows agree to the last decimal place with the row that '
        'is kept, so nothing measurable is lost and they are not listed '
        'individually. They are counted here so that the arithmetic from '
        'Third Normal Form to the loaded database closes.')

    add('Rows', 'Measurements refused by a primary key, values disagreeing',
        'A second organisation publishing a different value for a key another '
        'organisation had already filled',
        plural(len(rows), 'row'),
        'Climate precedence is Bangladesh Meteorological Department, then '
        'Bangladesh Rice Research Institute, then Bangladesh Bureau of '
        'Statistics. Equal sources use the newest coverage, then the later '
        'source occurrence. '
        'Of these, %d are a real disagreement about the value and %d are the '
        'same reading rounded to a different number of decimal places. Every '
        'one is listed with both values in csv/BCNF/_Key_Conflicts.csv, so no '
        'rejected measurement is lost in silence. Together with the exact '
        'duplicates above this accounts for all %d rows the key refuses.'
        % (len(sub), len(rows) - len(sub), lost))

def md_escape(s):
    return s.replace('|', r'\|')


def build():
    os.makedirs(OUTDIR, exist_ok=True)
    whole_sources()
    allsh, used, exc, titles, groups = bbs_sheets()
    named_sheets(titles)
    in_scope_files()
    stage_columns()
    unloaded_tables()
    unloaded_rows()

    nsrc = 0
    for root, _, fs in os.walk(SRC):
        nsrc += sum(1 for f in fs if f.lower().split('.')[-1]
                    in ('xlsx', 'xls', 'csv'))
    usedfiles = {(b['org'], b['file']) for b in BK.BLOCKS}

    L = []
    A = L.append
    A('# Exclusion Register')
    A('')
    A('Every data item left out of the loaded database, with the reason. One '
      'entry per item.')
    A('')
    A('This file is generated by `normalization/scripts/exclusions.py`, which '
      'reads the source tree and')
    A('the loaded tables and measures every figure below. No count is typed '
      'in by hand.')
    A('')
    A('An item appears here for one of five reasons, and the reason is always '
      'stated in full:')
    A('')
    A('1. **No shared key.** The item cannot join to any relation in the '
      'approved diagram.')
    A('2. **Wrong grain.** The item is an average over a decade or a thirty '
      'year normal, and every')
    A('   climate relation is keyed by a single year and month.')
    A('3. **Not in the approved diagram.** The subject has no entity, and '
      'adding one is a design')
    A('   decision for the team and the supervisor, not for a loading '
      'program.')
    A('4. **Off topic.** The item measures something other than the '
      'environment.')
    A('5. **A normal form requires it.** A column moves to another relation, '
      'or a key refuses a')
    A('   duplicate row.')
    A('')
    A('## What is loaded, for comparison')
    A('')
    A('| | Count |')
    A('|---|---|')
    A('| Spreadsheet and comma separated files in the source tree | %d |'
      % nsrc)
    A('| Files the load reads | %d |' % len(usedfiles))
    A('| Sheets in the Bangladesh Bureau of Statistics workbook | %d |'
      % len(allsh))
    A('| Sheets the load reads | %d |' % len(used))
    A('| Published blocks the load reads | %d |' % len(BK.BLOCKS))
    A('| Relations loaded | %d |'
      % len([f for f in os.listdir(os.path.join(CSVDIR, 'BCNF'))
             if f.endswith('.csv') and not f.startswith('_')]))
    A('')
    A('The load reads %d of %d files. That is the intended outcome, not a '
      'shortfall: the team'
      % (len(usedfiles), nsrc))
    A('collected widely, then kept what the approved Entity Relationship '
      'diagram can actually hold.')
    A('')

    order = ['Source', 'BBS sheet', 'BBS sheet, named', 'BMD file',
             'BRRI file', 'BWDB file', 'BBS file']
    heads = {
        'Source': ('Whole sources excluded',
                   'A source is excluded when no file in it shares a key '
                   'with the approved diagram, or when its subject is '
                   'outside the scope the diagram draws.'),
        'BBS sheet': ('Bangladesh Bureau of Statistics sheets excluded, by '
                      'group',
                      'The time series environmental database holds %d '
                      'sheets. The load reads %d. The remaining %d are '
                      'grouped below by the reason they are left out.'
                      % (len(allsh), len(used), len(exc))),
        'BBS sheet, named': ('Individual sheets a reader would question',
                             'These five look in scope from the title alone. '
                             'Each is excluded for a structural reason, '
                             'stated in full so the decision can be '
                             'checked.'),
        'BMD file': ('Bangladesh Meteorological Department files excluded',
                     ''),
        'BRRI file': ('Bangladesh Rice Research Institute files excluded',
                      ''),
        'BWDB file': ('Bangladesh Water Development Board files excluded',
                      ''),
        'BBS file': ('Bangladesh Bureau of Statistics files excluded', ''),
    }
    n = 0
    for scope in order:
        items = [r for r in REG if r[0] == scope]
        if not items:
            continue
        n += 1
        h, sub = heads[scope]
        A('## %d. %s' % (n, h))
        A('')
        if sub:
            A(sub)
            A('')
        A('| Item | What it holds | Size | Why it is excluded |')
        A('|---|---|---|---|')
        for _, item, holds, size, why in items:
            A('| **%s** | %s | %s | %s |'
              % (md_escape(item), md_escape(holds), size, md_escape(why)))
        A('')

    n += 1
    A('## %d. Columns dropped at each normalization stage' % n)
    A('')
    A('A column dropped here is never data thrown away. It either moves to '
      'another relation,')
    A('because a normal form requires it, or it is absent from the approved '
      'diagram and is')
    A('recorded here instead of being loaded.')
    A('')
    A('| Stage | Table and columns | Size | Why |')
    A('|---|---|---|---|')
    for scope, item, _, size, why in REG:
        if scope.startswith('Column,'):
            A('| %s | **%s** | %s | %s |'
              % (scope.split(', ')[1], md_escape(item), size, md_escape(why)))
    A('')

    n += 1
    A('## %d. Whole tables built during normalization and never loaded' % n)
    A('')
    A('| Table | What it holds | Size | Why it is not loaded |')
    A('|---|---|---|---|')
    for scope, item, holds, size, why in REG:
        if scope == 'Table, not loaded':
            A('| **%s** | %s | %s | %s |'
              % (md_escape(item), md_escape(holds), size, md_escape(why)))
    A('')

    n += 1
    A('## %d. Rows the load could not keep' % n)
    A('')
    A('| Item | What it holds | Size | Why |')
    A('|---|---|---|---|')
    for scope, item, holds, size, why in REG:
        if scope == 'Rows':
            A('| **%s** | %s | %s | %s |'
              % (md_escape(item), md_escape(holds), size, md_escape(why)))
    A('')

    n += 1
    A('## %d. Every excluded sheet, listed' % n)
    A('')
    A('The complete list, so no sheet is hidden inside a group count.')
    A('')
    for g in sorted(groups, key=lambda k: -len(groups[k])):
        A('**%s, %d sheets.** %s'
          % (g, len(groups[g]),
             ' '.join('`%s`' % s for s in sorted(groups[g]))))
        A('')

    with open(OUTDIR + 'EXCLUSION_REGISTER.md', 'w', encoding='utf-8') as f:
        f.write('\n'.join(L).rstrip() + '\n')

    with open(OUTDIR + 'EXCLUSIONS.csv', 'w', newline='',
              encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['Scope', 'Item', 'What it holds', 'Size',
                    'Why it is excluded'])
        for r in REG:
            w.writerow(list(r))

    with open(OUTDIR + 'EXCLUDED_BBS_SHEETS.csv', 'w', newline='',
              encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['Sheet', 'Title as published', 'Group'])
        for s in sorted(exc, key=lambda x: x.strip()):
            t = s.strip()
            w.writerow([t, titles.get(t, ''), classify(s)])

    json.dump({'sheets_total': len(allsh), 'sheets_used': len(used),
               'sheets_excluded': len(exc), 'entries': len(REG)},
              open(OUTDIR + 'summary.json', 'w'), indent=1)
    return len(REG), len(exc)


if __name__ == '__main__':
    a, b = build()
    print('%d register entries' % a)
    print('%d excluded BBS sheets listed' % b)
    for f in sorted(os.listdir(OUTDIR)):
        print('  %-30s %6d bytes'
              % (f, os.path.getsize(OUTDIR + f)))

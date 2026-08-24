"""0NF block inventory + verbatim 0NF CSV emission.

Every in-scope source block is declared here with its physically measured
boundaries. This module is imported by extract.py; it does not run alone.
"""
import openpyxl
import xlrd
import csv
import os

from nlib import (SRC, BBS_XLSX, BMD_TEMP, BMD_SUN, BRRI_DIR, BWDB_RIVERS,
                  BWDB_GW, write_raw_csv, clean_text, anom)

_CACHE = {}


def bbs_rows(sheet):
    """All rows of a BBS sheet, trailing all-blank rows trimmed."""
    key = ('bbs', sheet)
    if key not in _CACHE:
        wb = _CACHE.get('bbs_wb')
        if wb is None:
            wb = openpyxl.load_workbook(BBS_XLSX, read_only=True,
                                        data_only=True)
            _CACHE['bbs_wb'] = wb
        rs = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        while rs and all(c is None or clean_text(c) == '' for c in rs[-1]):
            rs.pop()
        _CACHE[key] = rs
    return _CACHE[key]


def bmd_temp_rows(sheet):
    key = ('bmdt', sheet)
    if key not in _CACHE:
        wb = _CACHE.get('bmdt_wb')
        if wb is None:
            wb = openpyxl.load_workbook(BMD_TEMP, read_only=True,
                                        data_only=True)
            _CACHE['bmdt_wb'] = wb
        rs = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        while rs and all(c is None or clean_text(c) == '' for c in rs[-1]):
            rs.pop()
        _CACHE[key] = rs
    return _CACHE[key]


def bmd_sun_rows(sheet):
    key = ('bmds', sheet)
    if key not in _CACHE:
        bk = _CACHE.get('bmds_wb')
        if bk is None:
            bk = xlrd.open_workbook(BMD_SUN)
            _CACHE['bmds_wb'] = bk
        sh = bk.sheet_by_name(sheet)
        rs = [[sh.cell_value(i, j) for j in range(sh.ncols)]
              for i in range(sh.nrows)]
        while rs and all(c is None or clean_text(c) == '' for c in rs[-1]):
            rs.pop()
        _CACHE[key] = rs
    return _CACHE[key]


BRRI_FILES = {
    'Max_Temp': 'BRRI_Daily_Maximum_Temperature.xlsx',
    'Min_Temp': 'BRRI_Daily_Minimum_Temperature.xlsx',
    'Rainfall': 'BRRI_Daily_Total_Rainfall.xlsx',
    'Humidity': 'BRRI_Daily_Average_Humidity.xlsx',
    'Sunshine': 'BRRI_Daily_Sunshine.xlsx',
}


def brri_rows(tag):
    key = ('brri', tag)
    if key not in _CACHE:
        wb = openpyxl.load_workbook(BRRI_DIR + BRRI_FILES[tag],
                                    read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rs = [list(r) for r in ws.iter_rows(values_only=True)]
        while rs and all(c is None or clean_text(c) == '' for c in rs[-1]):
            rs.pop()
        _CACHE[key] = rs
        wb.close()
    return _CACHE[key]


def bwdb_gw_rows(sheet):
    key = ('gw', sheet)
    if key not in _CACHE:
        wb = _CACHE.get('gw_wb')
        if wb is None:
            wb = openpyxl.load_workbook(BWDB_GW, read_only=True,
                                        data_only=True)
            _CACHE['gw_wb'] = wb
        rs = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
        while rs and all(c is None or clean_text(c) == '' for c in rs[-1]):
            rs.pop()
        _CACHE[key] = rs
    return _CACHE[key]


def rivers_rows():
    if 'rivers' not in _CACHE:
        with open(BWDB_RIVERS, encoding='utf-8-sig', newline='') as f:
            _CACHE['rivers'] = [r for r in csv.reader(f)]
    return _CACHE['rivers']


BBS_FILE = 'BBS_Time_Series_Environmental_Database.xlsx'
BMD_TEMP_FILE = 'Temperature Data.xlsx'

BLOCKS = []


def B(bid, org, fil, sheet, title, lo, hi, ncols, rep, hdr, target,
      loader=None, note=None):
    BLOCKS.append(dict(id=bid, org=org, file=fil, sheet=sheet, title=title,
                       lo=lo, hi=hi, ncols=ncols, rep=rep, hdr=hdr,
                       target=target, loader=loader or 'bbs', note=note))


B('B01', 'BBS', BBS_FILE, 'T1.02',
  'Monthly Average Minimum Temperature by Station, 2011-16 (Celsius)',
  0, 251, 15, '12 month columns Jan..Dec plus Annual',
  'r0 title, r1 Year+month names, r2 column-number row',
  'Temperature_Record (Type=Minimum)',
  note='Station identity is carried by "Station : X" label rows inside the '
       'data area, not by a column.')
B('B02', 'BBS', BBS_FILE, 'T1.02',
  'Monthly Average Minimum Temperature in celcius by Station, 2017-2024',
  252, 291, 98, '8 year blocks x 12 month columns (96 measure columns)',
  'r252 title, r253 Station+year block heads, r254 month names, '
  'r255 column-number row',
  'Temperature_Record (Type=Minimum)')

B('B03', 'BBS', BBS_FILE, 'T1.03',
  'Monthly Average Maximum Temperature by Station, 2011-16 (Celsius)',
  0, 253, 15, '12 month columns Jan..Dec plus Annual',
  'r0 title, r1 Year+month names, r2 column-number row',
  'Temperature_Record (Type=Maximum)',
  note='Station identity is carried by "Station : X" label rows.')
B('B04', 'BBS', BBS_FILE, 'T1.03',
  'Monthly Average Maximum Temperature in celcius by Station, 2017-2024',
  254, 293, 98, '8 year blocks x 12 month columns',
  'r254 title, r255 Station+year block heads, r256 month names, '
  'r257 column-number row',
  'Temperature_Record (Type=Maximum)')

B('B05', 'BBS', BBS_FILE, 'T1.05',
  'Monthly Total Rainfall by Center/Station, 2011-14 (MM)',
  0, 146, 16, '12 month columns Jan..Dec plus Annual',
  'r0 title, r1 Station/Year/month names, r2 column-number row',
  'Rainfall_Record')
B('B06', 'BBS', BBS_FILE, 'T1.05',
  'Monthly and Yearly Total Rainfall data in millimeter by Station, 2015-2019',
  147, 328, 16, '12 month columns Jan..Dec plus Total',
  'r147 title, r148 header, r149 column-number row',
  'Rainfall_Record')
B('B07', 'BBS', BBS_FILE, 'T1.05',
  'Monthly and Yearly Total Rainfall data in Millimeter by Station, 2020-2024',
  329, 510, 16, '12 month columns Jan..Dec plus Total',
  'r329 title, r330 header (Name of Station), r331 column-number row',
  'Rainfall_Record')

B('B08', 'BBS', BBS_FILE, 'T1.06',
  'Monthly & Yearly Average Humidity in Percent by Station, 2011-14',
  0, 150, 16, '12 month columns Jan..Dec plus Annual',
  'r0 title, r1 header (Station name), r2 column-number row',
  'Humidity_Record')
B('B09', 'BBS', BBS_FILE, 'T1.06',
  'Monthly & Yearly Average Humidity in percent by Station, 2015-2019',
  151, 332, 16, '12 month columns Jan..Dec plus Annual',
  'r151 title, r152 header, r153 column-number row',
  'Humidity_Record')
B('B10', 'BBS', BBS_FILE, 'T1.06',
  'Monthly & Yearly Average Humidity in percent by Station, 2020-2024',
  333, 515, 16, '12 month columns Jan..Dec plus Annual',
  'r334 title, r335 header (Name of Station), r336 column-number row',
  'Humidity_Record')

B('B11', 'BBS', BBS_FILE, 'T1.08',
  'Monthly Minimum Wind speed in Knots and Directioction in Degrees by '
  'Station, 2015-2024',
  0, 393, 26, '12 month groups, each a (spd, dir) column pair',
  'r0 title, r1 Station/Year+month names, r2 spd/dir subheader, '
  'r3 column-number row',
  'Wind_Record (Type=Minimum)',
  note='Station identity is carried by "Station : X" label rows.')

B('B12', 'BBS', BBS_FILE, 'T1.10',
  'Monthly Highest Maximun Wind Speed by Station, 2011-16 (Data in Knots)',
  0, 244, 26, '12 month groups, each a (SPD., DIR.) column pair',
  'r0 title, r1 Year+month names, r2 SPD./DIR. subheader, '
  'r3 column-number row',
  'Wind_Record (Type=Maximum)',
  note='Station label rows also carry Lat./Long. strings in mid-row columns.')
B('B13', 'BBS', BBS_FILE, 'T1.10',
  'Monthly Highest Maximun Wind Speed in Knots by Station, 2017-2024',
  245, 572, 26, '12 month groups, each a (SPD., DIR.) column pair',
  'r245 title, r246 Year+month names, r247 SPD./DIR. subheader, '
  'r248 column-number row',
  'Wind_Record (Type=Maximum)')

B('B14', 'BBS', BBS_FILE, 'T1.13',
  'Daily Total Sun Shine Hours by Station, 2019-2024',
  0, 2525, 35, '31 day columns (Day 1..31)',
  'r0 title, r1 Stations/Year/Month/Day, r2 day numbers, '
  'r3 column-number row',
  'Sunshine_Record')

for _bid, _st, _lo, _hi, _t in (
        ('B15', 'Bogura', 0, 186, 'Radiation BY Station (Bogura), 2019-2022'),
        ('B16', 'Dinajpur', 186, 375,
         'Radiation BY Station (Dinajpur), 2019-2022'),
        ('B17', 'Dhaka', 375, 562, 'Radiation BY Station (Dhaka), 2019-2022'),
        ('B18', 'Satkhira', 562, 749,
         'Radiation BY Station (Satkhira), 2019-2022'),
        ('B19', 'Chuadanga', 749, 936,
         'Radiation BY Station (Chuadanga), 2019-2022')):
    B(_bid, 'BBS', BBS_FILE, 'T1.14', _t, _lo, _hi, 190,
      '6 year blocks (2019..2024) x 31 day columns D_01..D_31',
      'title row, header row (Station Name/Month/Date + year heads), '
      'D_01..D_31 subheader row, column-number row',
      'Radiation_Record',
      note='Column labelled "Date" is the hour of observation (0..13), not '
           'the day of month. Blank separator rows sit between months.')

B('B20', 'BBS', BBS_FILE, 'T1.19',
  'Monthly Frequency of Thunderstorm by Station, 2015-2024',
  0, 354, 15, '12 month columns Jan..Dec',
  'r0 title, r1 Station/Year+month names, r2 column-number row',
  'Climatic_Event_Record.Thunderstorm')
B('B21', 'BBS', BBS_FILE, 'T1.19 a',
  'Monthly Frequency of Lightening by Station, 2015-2024',
  0, 354, 15, '12 month columns Jan..Dec',
  'r0 title, r1 Station/Year+month names, r2 column-number row',
  'Climatic_Event_Record.Lightning')

B('B22', 'BBS', BBS_FILE, 'T1.66',
  'District Wise Total Area of Forest from 2019-20 to 2023-24 (In Acre)',
  0, 42, 43, '5 fiscal-year blocks x 8 measure columns',
  'r0 title, r1 SL/District+fiscal years, r2 measure labels, '
  'r3 Section 20 / Section 4&6 sublabels, r4 column-number row',
  'Forest_Area_Record')

_T84 = (('B23', 'T1.84', 0, 182, 'Biochemical oxygen demand (BOD)'),
        ('B25', 'T1.84b', 0, 182, 'Chemical Oxygen Demand (COD)'),
        ('B26', 'T1.84c', 0, 182, 'pH/Acidity/Alkalinity'),
        ('B27', 'T184d', 0, 199, 'Salinity'),
        ('B28', 'T1.84e', 0, 180, 'Dissolved Oxygen (DO)'))
for _bid, _sh, _lo, _hi, _par in _T84:
    B(_bid, 'BBS', BBS_FILE, _sh,
      'Freshwater Quality by Station, 2018-2023 (%s)' % _par,
      _lo, _hi, 17,
      'year blocks side by side, each repeating the District/Station column',
      'title row, year head row, parameter-category rows, '
      'column-number row',
      'Water_Quality + River + River_Station',
      note='River names appear as group header rows with no measurement '
           'value; station rows follow beneath them.')
B('B24', 'BBS', BBS_FILE, 'T1.84',
  'Freshwater Quality by District/Station, 2020-2023 (BOD, lakes)',
  183, 207, 8, '4 year columns 2020..2023',
  'r184 title, r185 header, r186/r187 category+parameter rows, '
  'r188 column-number row',
  'Water_Quality + River + River_Station',
  note='Second, independently headed block appended below the river block of '
       'the same sheet. Lake names act as group headers.')

_T85 = (('B29', 'T1.85', 'Biochemical oxygen demand (BOD)', 15),
        ('B30', 'T1.85b', 'Chemical oxygen demand (COD)', 15),
        ('B31', 'T1.85c', 'pH/Acidity/Alkalinity', 15),
        ('B32', 'T1.85d', 'Salinity', 15),
        ('B33', 'T1.85e', 'Dissolved oxygen (DO)', 15),
        ('B34', 'T1.85f', 'Plastic waste and other marine debris', 17))
for _bid, _sh, _par, _hi in _T85:
    B(_bid, 'BBS', BBS_FILE, _sh,
      'Marine Water Quality by District/Station, 2019-2024 (%s)' % _par,
      0, _hi, 8, '6 year columns 2019..2024',
      'r0 title, r1 District/Station+years, r2 category row, '
      'r3 parameter row, r4 column-number row',
      'Water_Quality')

B('B35', 'BBS', BBS_FILE, 'T3.14',
  'Treatment of Industrial Waste Water by type of Treatment (Cubic Meter) '
  '- Industry Type block',
  0, 33, 8, '3 fiscal-year quantity columns + 3 percentage columns',
  'r0 title, r1 header + Percentage span, r2 fiscal years under Percentage, '
  'r3 column-number row',
  'Type_Of_Establishments',
  note='Only the "Industry Type" sub-block (Micro/Small/Medium/Large) is in '
       'scope; the sheet also carries treatment-type and category-type '
       'blocks.')
B('B36', 'BBS', BBS_FILE, 'T3.22',
  'Distribution of Reused Waste Water by Type and Category of Industry '
  '(000 litre.)',
  0, 7, 8, '3 fiscal-year produced columns + 3 reuse columns',
  'r0 title, r1 header spans, r2 fiscal years, r3 column-number row',
  'Industrial_Type')

for _i in range(1, 19):
    B('B%02d' % (36 + _i), 'BMD', BMD_TEMP_FILE, 'Table-%d' % _i,
      'Monthly Temperature (degC) in %d' % (1994 + _i),
      0, 36, 25, 'Max_Jan..Max_Dec and Min_Jan..Min_Dec (24 columns)',
      'r0 title, r1 header',
      'Temperature_Record (both Types)', loader='bmdt')

B('B55', 'BMD', 'Sunshine.xls', 'Climate ',
  'Climate (daily sunshine grid, 2022)',
  0, 17, 34, '31 day columns D_01..D_31', 'r0 header',
  'Sunshine_Record', loader='bmds',
  note='Station ID is 0.0 on every row and no station name is present.')
B('B56', 'BMD', 'Sunshine.xls', 'Climate Leap year February',
  'Climate Leap year February (daily sunshine grid, 2024)',
  0, 2, 34, '31 day columns D_01..D_31', 'r0 header',
  'Sunshine_Record', loader='bmds',
  note='Station ID is 0.0 and no station name is present.')

_BRRI = (('B57', 'Max_Temp', 'Daily Maximum Temperature data in degree celcius',
          'Temperature_Record (Type=Maximum, monthly MEAN)'),
         ('B58', 'Min_Temp', 'Daily Minimum Temperature data in celcius',
          'Temperature_Record (Type=Minimum, monthly MEAN)'),
         ('B59', 'Rainfall', 'Daily total Rain fall data in Milli meter',
          'Rainfall_Record (monthly SUM)'),
         ('B60', 'Humidity', 'Daily average Humidity in %',
          'Humidity_Record (monthly MEAN)'),
         ('B61', 'Sunshine', 'Daily Total Sunshine Hours',
          'Sunshine_Record (daily, no aggregation)'))
for _bid, _tag, _t, _tgt in _BRRI:
    B(_bid, 'BRRI', BRRI_FILES[_tag], BRRI_FILES[_tag].replace('.xlsx', ''),
      _t, 0, None, 34, '31 day columns (Days 1..31)',
      'r0 Stations/Year/Month/Days, r1 day numbers', _tgt, loader='brri')

B('B62', 'BWDB', 'BWDB_Rivers_Information.csv', '(csv)',
  'BWDB river register', 0, None, 5, 'none (already one row per river)',
  'row 0 header', 'River', loader='rivers')
for _i, _sh in enumerate(('Table-2', 'Table-3', 'Table-4', 'Table-5',
                          'Table-6', 'Table-7')):
    B('B%d' % (63 + _i), 'BWDB',
      'BWDB_Groundwater_Weekly_Data_2018.xlsx', _sh,
      'Groundwater observation wells (%s)' % _sh, 0, None, 7,
      'none (one row per well)', 'r0 title, r1 header',
      'Ground_Water_Well', loader='gw')


def block_rows(b):
    """Raw rows of a block, exactly as published."""
    ld = b['loader']
    if ld == 'bbs':
        rs = bbs_rows(b['sheet'])
    elif ld == 'bmdt':
        rs = bmd_temp_rows(b['sheet'])
    elif ld == 'bmds':
        rs = bmd_sun_rows(b['sheet'])
    elif ld == 'brri':
        tag = [k for k, v in BRRI_FILES.items() if v == b['file']][0]
        rs = brri_rows(tag)
    elif ld == 'gw':
        rs = bwdb_gw_rows(b['sheet'])
    elif ld == 'rivers':
        rs = rivers_rows()
    else:
        raise ValueError(ld)
    hi = b['hi'] if b['hi'] is not None else len(rs)
    return rs[b['lo']:min(hi, len(rs))]


def emit_0nf():
    """Write one verbatim CSV per raw block. Returns the measured inventory."""
    inv = []
    for b in BLOCKS:
        rs = block_rows(b)
        ncols = max([len(r) for r in rs] + [0])
        ncols = min(ncols, b['ncols']) if b['ncols'] else ncols
        real = 0
        for r in rs:
            for j in range(len(r) - 1, -1, -1):
                if r[j] is not None and clean_text(r[j]) != '':
                    real = max(real, j + 1)
                    break
        ncols = max(ncols, real) if real else ncols
        name = '%s__%s__%s' % (b['org'],
                               os.path.splitext(b['file'])[0]
                               .replace(' ', '_'),
                               b['sheet'].strip().replace(' ', '_')
                               .replace('/', '-') or 'sheet')
        name = '%s__%s' % (b['id'], name)
        full, written = write_raw_csv(name, rs, ncols, note=b['note'])
        inv.append(dict(b, csv=name + '.csv', raw_rows=full,
                        raw_cols=ncols, written=written))
    return inv

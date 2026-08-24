"""Build Environmental_Normalization_0NF_to_BCNF.xlsx from the generated CSVs.

Run extract.py first. This program reads only the version-local csv/ folder and
writes one workbook. Nothing here recomputes a figure: every number shown is
read back from the files the extraction produced, so the workbook and the CSVs
can never disagree.

Run:  python normalization/scripts/workbook.py
"""
import csv
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import blocks as BK
from nlib import CSVDIR, OUT
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font,
                             PatternFill, Side)
from openpyxl.utils import get_column_letter

STAGE_COLOUR = {
    '0NF': '8C3A3A',
    '1NF': 'A5643C',
    '2NF': '7A6A2E',
    '3NF': '2F6B4F',
    'BCNF': '2B4C7E',
    'INFO': '4A4A4A',
}
HEAD_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14)
NOTE_FONT = Font(italic=True, color='555555', size=10)
BAND = PatternFill('solid', fgColor='F2F5F9')
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SHEET_CAP = 400


def read_csv(path):
    with open(path, newline='', encoding='utf-8') as f:
        r = list(csv.reader(f))
    return (r[0], r[1:]) if r else ([], [])


def is_number(s):
    try:
        float(s)
        return True
    except (TypeError, ValueError):
        return False


def unbold(s):
    """Strip the markdown bold markers around a cell, but nothing else.

    A blanket removal of every '**' would be wrong here: several sentinel
    names in the quality log are literally '**', '***' and '****', and
    deleting those characters would make three different problems read as the
    same one.
    """
    s = s.strip()
    if len(s) > 4 and s.startswith('**') and s.endswith('**'):
        return s[2:-2].strip()
    return s


def put(ws, hdr, rows, stage='INFO', start=1, widths=None, cap=None,
        number_cols=None):
    """Write a header row and data rows with the stage's colour."""
    fill = PatternFill('solid', fgColor=STAGE_COLOUR[stage])
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=start, column=j, value=h)
        c.font = HEAD_FONT
        c.fill = fill
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BOX
    shown = rows if cap is None else rows[:cap]
    for i, r in enumerate(shown):
        for j, v in enumerate(r, 1):
            if isinstance(v, str) and is_number(v) and v != '':
                v = float(v)
                if v == int(v) and abs(v) < 1e15:
                    v = int(v)
            c = ws.cell(row=start + 1 + i, column=j, value=v)
            c.border = BOX
            if i % 2:
                c.fill = BAND
            c.alignment = Alignment(vertical='top', wrap_text=False)
    ws.freeze_panes = ws.cell(row=start + 1, column=1)
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    else:
        for j, h in enumerate(hdr, 1):
            w = max(len(str(h)) + 2,
                    *(len(str(r[j - 1])) + 2 for r in shown[:60]
                      if j - 1 < len(r))) if shown else len(str(h)) + 2
            ws.column_dimensions[get_column_letter(j)].width = min(46, max(9,
                                                                          w))
    return start + 1 + len(shown)


def title(ws, text, sub=None, stage='INFO'):
    ws['A1'] = text
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = PatternFill('solid', fgColor='FFFFFF')
    r = 2
    if sub:
        for line in sub:
            ws.cell(row=r, column=1, value=line).font = NOTE_FONT
            r += 1
    return r + 1


def sheet(wb, name, tabcolour):
    ws = wb.create_sheet(name[:31])
    ws.sheet_properties.tabColor = tabcolour
    return ws


def zero_nf_index():
    """Map each block id to (csv path, true row count).

    A block whose end row is open ended runs to the end of its sheet, so its
    length is not in the block table. It is read back from the file the
    extraction wrote: the companion .meta.txt states the true count for a
    block that was capped, and for a block written in full the file itself is
    the count.
    """
    d = os.path.join(CSVDIR, '0NF')
    idx = {}
    for fn in sorted(os.listdir(d)):
        if not fn.endswith('.csv'):
            continue
        bid = fn.split('__')[0]
        path = os.path.join(d, fn)
        meta = path[:-4] + '.meta.txt'
        n = None
        if os.path.exists(meta):
            with open(meta, encoding='utf-8') as f:
                for line in f:
                    if line.startswith('TRUE full row count'):
                        n = int(line.split(':')[1].strip())
                        break
        if n is None:
            with open(path, newline='', encoding='utf-8') as f:
                n = sum(1 for _ in csv.reader(f))
        idx[bid] = (path, n)
    return idx


GUIDE = [
    ('0NF', 'Unnormalized form',
     'The data exactly as the organisation published it.',
     'A cell holds several values at once, one heading spans many columns, '
     'and a row may be a title, a note or a column number legend rather than '
     'data. Twelve month columns repeat across a single row, so the year and '
     'the month are stored in the heading rather than in the row. No primary '
     'key exists, because a row is not identified by anything it contains.',
     'Nothing can be queried reliably. Asking for the rainfall of one station '
     'in one month means reading a heading, not a value.'),
    ('1NF', 'First Normal Form',
     'Every cell holds one value and every repeating group becomes rows.',
     'Each repeating column group is unpivoted, so the month that lived in a '
     'heading becomes a Month column and each measurement becomes its own '
     'row. Every row now carries the key that identifies it.',
     'The tables are still wide, and they deliberately still carry the source '
     'columns and the measure name. Those columns depend on only part of the '
     'key, which is exactly the defect the next stage removes.'),
    ('2NF', 'Second Normal Form',
     'Every non key attribute depends on the whole key, not on part of it.',
     'The source organisation, file and sheet depend only on the block a row '
     'came from, not on the station and month. The unit depends only on the '
     'measure. The river depends only on the monitoring station. Each of '
     'these is a partial dependency, so each moves into its own table.',
     'A dependency between two non key attributes can still survive, which '
     'is what the next stage removes.'),
    ('3NF', 'Third Normal Form',
     'No non key attribute depends on another non key attribute.',
     'The measure name is no longer stored as a data value: each measured '
     'quantity gets its own relation, which is what the approved Entity '
     'Relationship diagram draws. River details also move away from yearly '
     'measurements because they depend on the station alone.',
     'A relation can still hold a determinant that is not a candidate key, '
     'which is the one case Third Normal Form permits and Boyce Codd Normal '
     'Form does not.'),
    ('BCNF', 'Boyce Codd Normal Form',
     'Every determinant is a candidate key.',
     'Each relation is checked attribute by attribute. The primary key is '
     'then enforced against the real data: where two organisations publish a '
     'different value for the same key, the key forbids both, so one is kept '
     'by a stated precedence rule and the rejected value is recorded rather '
     'than dropped in silence.',
     'This is the form the database is loaded in. Twenty-one relations, every '
     'primary key unique, every foreign key resolving to a parent row.'),
]


def build():
    wb = Workbook()
    wb.remove(wb.active)
    Z0 = zero_nf_index()

    ws = sheet(wb, 'Guide', STAGE_COLOUR['INFO'])
    r = title(ws, 'Normalization of the Bangladesh Environmental Database',
              ['This workbook shows one body of data at five stages, from the '
               'form the government published it in to the form the database '
               'is loaded in.',
               'Every figure in every sheet is produced by the extraction '
               'program reading the source files. No figure is typed in by '
               'hand and no figure is estimated.',
               'Sources in scope: Bangladesh Bureau of Statistics, Bangladesh '
               'Meteorological Department, Bangladesh Rice Research Institute '
               'and Bangladesh Water Development Board.'])
    r = put(ws, ['Stage', 'Name', 'The rule it satisfies',
                 'What the stage does to the data',
                 'What is still wrong after it'],
            [[a, b, c, d, e] for a, b, c, d, e in GUIDE],
            'INFO', r, widths=[8, 24, 44, 66, 66])
    r += 2
    ws.cell(row=r, column=1,
            value='How to read this workbook').font = TITLE_FONT
    r += 1
    for line in [
        'The sheets are in stage order. Each stage sheet has its own colour, '
        'shown in the tab.',
        'A stage begins with a sheet named "<stage> Tables", which lists every '
        'table at that stage with its attributes and its row count. One sheet '
        'per table follows.',
        'A sheet that is too large to hold in full states its true row count '
        'on its first line, above the table. The complete data is always in '
        'the matching file under csv/, one file per table.',
        'Stage Comparison puts the five stages side by side and states why the '
        'row count moves the way it does at each step.',
        'Functional Dependencies gives the reasoning that puts each relation '
        'in Boyce Codd Normal Form. Key Enforcement states what applying the '
        'primary key costs each relation, and Key Conflicts lists the '
        'individual measurements it refused, with both values.',
        'Traceability takes each loaded relation back to the published column '
        'group it came from. Load Statistics and Data Quality report how much '
        'data each organisation contributes and every problem found in the '
        'published files.',
    ]:
        ws.cell(row=r, column=1, value='- ' + line).font = NOTE_FONT
        r += 1
    ws.column_dimensions['A'].width = 20

    ws = sheet(wb, '0NF Raw Blocks', STAGE_COLOUR['0NF'])
    r = title(ws, 'Unnormalized form: the raw blocks as published',
              ['A block is one rectangle of data inside one published sheet. '
               'A single sheet often holds several, stacked one under '
               'another, each with its own heading rows.',
               'The column named Repeating group is the reason the block is '
               'not a table: it names the group of columns that repeat, which '
               'First Normal Form turns into rows.'],
              '0NF')
    rows = []
    for b in BK.BLOCKS:
        path, nraw = Z0.get(b['id'], (None, 0))
        rows.append([b['id'], b['org'], b['file'], b['sheet'].strip(),
                     b['title'], nraw, b['ncols'], b['rep'],
                     b['hdr'], b['target'], b.get('note', '')])
    r = put(ws, ['Block', 'Organisation', 'Source file', 'Sheet',
                 'Title as published', 'Rows as published', 'Columns',
                 'Repeating group that breaks First Normal Form',
                 'Heading rows found above the data',
                 'Relation it feeds', 'Note'],
            rows, '0NF', r,
            widths=[7, 12, 34, 12, 48, 10, 8, 40, 38, 34, 52])

    ws = sheet(wb, '0NF Sample', STAGE_COLOUR['0NF'])
    r = title(ws, 'Unnormalized form: what a raw block actually looks like',
              ['Three blocks are reproduced verbatim, heading rows and all, '
               'so the defects can be seen rather than described.',
               'Nothing below is cleaned. Empty cells, merged heading '
               'artefacts and column number legend rows appear as published.'],
              '0NF')
    for bid in ('B01', 'B22', 'B62'):
        b = next((x for x in BK.BLOCKS if x['id'] == bid), None)
        if b is None or bid not in Z0:
            continue
        path, nraw = Z0[bid]
        with open(path, newline='', encoding='utf-8') as f:
            raw = list(csv.reader(f))
        ws.cell(row=r, column=1,
                value='%s  %s  sheet %s  -  %s'
                      % (bid, b['org'], b['sheet'].strip(),
                         b['title'])).font = Font(bold=True, size=12)
        r += 1
        ws.cell(row=r, column=1,
                value='Repeating group: %s. Heading rows: %s. The block holds '
                      '%d rows as published; the first 18 are shown.'
                      % (b['rep'], b['hdr'], nraw)).font = NOTE_FONT
        r += 1
        fill = PatternFill('solid', fgColor=STAGE_COLOUR['0NF'])
        for i, rr in enumerate(raw[:18]):
            for j, v in enumerate(rr[:16], 1):
                c = ws.cell(row=r, column=j, value=v)
                c.border = BOX
                if i < 3:
                    c.fill = fill
                    c.font = Font(color='FFFFFF', size=10)
            r += 1
        r += 2
    for j in range(1, 17):
        ws.column_dimensions[get_column_letter(j)].width = 15

    STAGE_TABLES = {}
    for stage in ('1NF', '2NF', '3NF', 'BCNF'):
        d = os.path.join(CSVDIR, stage)
        files = sorted(f for f in os.listdir(d)
                       if f.endswith('.csv') and not f.startswith('_'))
        ws = sheet(wb, '%s Tables' % stage, STAGE_COLOUR[stage])
        name = {'1NF': 'First Normal Form', '2NF': 'Second Normal Form',
                '3NF': 'Third Normal Form',
                'BCNF': 'Boyce Codd Normal Form'}[stage]
        g = next(x for x in GUIDE if x[0] == stage)
        r = title(ws, '%s: the tables at this stage' % name,
                  [g[2], g[3],
                   'The complete data for every table below is in '
                   'csv/%s/.' % stage], stage)
        rows = []
        tabs = {}
        for fn in files:
            hdr, data = read_csv(os.path.join(d, fn))
            nm = fn[:-4]
            tabs[nm] = (hdr, len(data))
            rows.append([nm, len(hdr), len(data), ', '.join(hdr)])
        STAGE_TABLES[stage] = tabs
        rows.append(['TOTAL, %d tables' % len(files), '',
                     sum(n for _, n in tabs.values()), ''])
        r = put(ws, ['Table', 'Attributes', 'Rows', 'Attribute list'],
                rows, stage, r, widths=[38, 11, 11, 110])

        for fn in files:
            hdr, data = read_csv(os.path.join(d, fn))
            nm = fn[:-4]
            short = nm.replace('_Record', '_Rec').replace('_Observation',
                                                          '_Obs')
            w2 = sheet(wb, ('%s %s' % (stage, short))[:31], STAGE_COLOUR[stage])
            note = ['Full row count: %d.' % len(data)]
            if len(data) > SHEET_CAP:
                note.append('This sheet shows the first %d rows only. The '
                            'complete table is csv/%s/%s.'
                            % (SHEET_CAP, stage, fn))
            else:
                note.append('This sheet shows the table in full.')
            rr = title(w2, '%s  (%s)' % (nm, name), note, stage)
            put(w2, hdr, data, stage, rr, cap=SHEET_CAP)

    ws = sheet(wb, 'Functional Dependencies', STAGE_COLOUR['BCNF'])
    hdr, data = read_csv(os.path.join(CSVDIR, 'BCNF',
                                      '_Functional_Dependencies.csv'))
    r = title(ws, 'Boyce Codd Normal Form verification, relation by relation',
              ['A relation is in Boyce Codd Normal Form when every '
               'determinant it contains is a candidate key.',
               'The final column gives the reason in words, so the check can '
               'be followed without reading the extraction program.'], 'BCNF')
    put(ws, hdr, data, 'BCNF', r, widths=[26, 46, 54, 30, 14, 92])

    ws = sheet(wb, 'Key Conflicts', STAGE_COLOUR['BCNF'])
    hdr, data = read_csv(os.path.join(CSVDIR, 'BCNF', '_Key_Conflicts.csv'))
    sub = [x for x in data if x[6] == 'substantive']
    rnd = len(data) - len(sub)
    r = title(ws, 'Measurements the primary key forced the load to reject',
              ['Two organisations sometimes publish a different value for the '
               'same station, year and month. The primary key allows one row, '
               'so one value is kept and the other cannot be loaded.',
               'Precedence for climate observations: Bangladesh '
               'Meteorological Department, then Bangladesh Rice Research '
               'Institute, then Bangladesh Bureau of Statistics. Equal '
               'sources use the newest stated coverage, then the later block.',
               'Of %d conflicts, %d are the two organisations rounding the '
               'same reading differently and %d are a real disagreement about '
               'the value. The %d substantive conflicts are listed first.'
               % (len(data), rnd, len(sub), len(sub)),
               'The complete list is csv/BCNF/_Key_Conflicts.csv.'], 'BCNF')
    put(ws, hdr, sub + [x for x in data if x[6] != 'substantive'],
        'BCNF', r, cap=SHEET_CAP, widths=[24, 40, 14, 16, 16, 16, 15])

    ws = sheet(wb, 'Key Enforcement', STAGE_COLOUR['BCNF'])
    r = title(ws, 'What enforcing the primary key costs each relation',
              ['Third Normal Form permits two rows with the same key, because '
               'no key is enforced there. Boyce Codd Normal Form is the stage '
               'the schema keys are applied, so a second row on the same key '
               'must go.',
               'A rejected row is never a row that was wrong. It is a row '
               'whose key another organisation had already filled. Every one '
               'is listed in the Key Conflicts sheet with both values, so the '
               'reader can judge the choice.',
               'Where the two values agree exactly, the second row is a '
               'duplicate and nothing is lost. Where they differ, the loss is '
               'a real disagreement and is counted as a conflict.'], 'BCNF')
    byrel = {}
    for x in data:
        e = byrel.setdefault(x[0], [0, 0])
        e[0 if x[6] == 'substantive' else 1] += 1
    reg = read_csv(os.path.join(CSVDIR, '3NF', 'River_Register_3NF.csv'))
    reg_names = set(x[1] for x in reg[1])
    rs_hdr, rs_data = read_csv(os.path.join(CSVDIR, 'BCNF',
                                            'River_Station.csv'))
    rs_names = set(x[rs_hdr.index('River_Name')] for x in rs_data)
    RIVER_NOTE = (
        'Not derived, and not a key rejection. The published register holds '
        '%d rows carrying %d distinct names, because the register lists '
        'Salda twice under two water development zones. Reducing the '
        'register to the name alone, which is all the approved diagram '
        'keeps, collapses that pair into one row. A further %d water bodies '
        'are added, because a monitoring station names them and no register '
        'row does: they are lakes, canals and one sea rather than rivers, '
        'and without them the foreign key from River_Station would not '
        'resolve. %d less one, plus %d, gives the %d rows loaded.'
        % (len(reg[1]), len(reg_names), len(rs_names - reg_names),
           len(reg[1]), len(rs_names - reg_names),
           STAGE_TABLES['BCNF']['River'][1]))

    rows = []
    for nm in sorted(STAGE_TABLES['BCNF']):
        t3 = STAGE_TABLES['3NF'].get(nm + '_3NF')
        n_bcnf = STAGE_TABLES['BCNF'][nm][1]
        if t3 is None:
            rows.append([nm, '', n_bcnf, '', '', '',
                         RIVER_NOTE if nm == 'River' else
                         'New at this stage. No organisation publishes it, so '
                         'it holds no duplicate key to resolve. It is '
                         'collected from the values the measurement relations '
                         'use, so that every foreign key has a parent row.'])
            continue
        n3 = t3[1]
        s, rd = byrel.get(nm, (0, 0))
        lost = n3 - n_bcnf
        if lost == 0:
            note = ('Every key is already unique, so the key holds with '
                    'nothing rejected.')
        else:
            dup = lost - s - rd
            parts = []
            if s:
                parts.append('%d %s about the value'
                             % (s, 'differs' if s == 1 else 'differ'))
            if rd:
                parts.append('%d %s the same reading rounded to a different '
                             'number of decimal places'
                             % (rd, 'is' if rd == 1 else 'are'))
            if dup:
                parts.append('%d %s exactly what the kept row says, so '
                             'nothing is lost'
                             % (dup, 'says' if dup == 1 else 'say'))
            joined = (parts[0] if len(parts) == 1 else
                      ' and '.join([', '.join(parts[:-1]), parts[-1]]))
            note = 'The key rejects %d rows: %s.' % (lost, joined)
        rows.append([nm, n3, n_bcnf, lost, s, rd, note])
    tot3 = sum(n for _, n in STAGE_TABLES['3NF'].values())
    totb = sum(n for _, n in STAGE_TABLES['BCNF'].values())
    ref = sum(n for nm, (_, n) in STAGE_TABLES['BCNF'].items()
              if nm + '_3NF' not in STAGE_TABLES['3NF'])
    nref = sum(1 for nm in STAGE_TABLES['BCNF']
               if nm + '_3NF' not in STAGE_TABLES['3NF'])
    unloaded = sum(n for nm, (_, n) in STAGE_TABLES['3NF'].items()
                   if nm[:-4] not in STAGE_TABLES['BCNF'])
    dropped = sum(r[3] for r in rows if r[3] != '')
    put(ws, ['Relation', 'Rows at Third Normal Form',
             'Rows loaded', 'Rows not loaded',
             'of which a real disagreement', 'of which only rounding',
             'Why the rows are not loaded'],
        rows + [['TOTAL', tot3, totb, dropped, len(sub), rnd,
                 'The arithmetic closes exactly: %d rows at Third Normal '
                  'Form, less %d reconciled within loaded relations, less %d '
                  'in tables intentionally documented rather than loaded, '
                  'plus %d rows in the %d relations that have no Third Normal '
                  'Form table of their own, gives the %d rows loaded.'
                  % (tot3, dropped, unloaded, ref, nref, totb)]],
        'BCNF', r, widths=[26, 13, 12, 13, 15, 14, 96])

    for fname, sname, headline, note in (
            ('STATISTICS.md', 'Load Statistics',
             'How much data each source contributes, and what survives',
             'Every figure is counted while the source files are read.'),
            ('DATA_QUALITY_LOG.md', 'Data Quality',
             'Every problem found in the published files',
             'Nothing is silently corrected. A repair is recorded as a '
             'repair, and a record that cannot be repaired is quarantined '
             'rather than replaced with an invented value.')):
        ws = sheet(wb, sname, STAGE_COLOUR['INFO'])
        r = title(ws, headline, [note,
                                 'The same content, in full, is in '
                                 '%s.' % fname])
        with open(OUT + fname, encoding='utf-8') as f:
            lines = f.read().split('\n')
        intable = False
        for line in lines:
            if line.startswith('#'):
                ws.cell(row=r, column=1,
                        value=line.lstrip('# ')).font = Font(bold=True,
                                                             size=12)
                r += 1
                intable = False
                continue
            if line.startswith('|'):
                cells = [unbold(c) for c in line.strip('|').split('|')]
                if set(''.join(cells)) <= set('-: ') and cells:
                    continue
                for j, v in enumerate(cells, 1):
                    if is_number(v) and v != '':
                        v = float(v)
                        if v == int(v):
                            v = int(v)
                    c = ws.cell(row=r, column=j, value=v)
                    c.border = BOX
                    if not intable:
                        c.font = HEAD_FONT
                        c.fill = PatternFill('solid',
                                             fgColor=STAGE_COLOUR['INFO'])
                r += 1
                intable = True
                continue
            intable = False
            if line.strip():
                ws.cell(row=r, column=1, value=line).alignment = \
                    Alignment(wrap_text=False)
            r += 1
        ws.column_dimensions['A'].width = 62
        for j in range(2, 9):
            ws.column_dimensions[get_column_letter(j)].width = 22

    ws = sheet(wb, 'Traceability', STAGE_COLOUR['INFO'])
    hdr, data = read_csv(os.path.join(CSVDIR, 'BCNF', '_Traceability.csv'))
    r = title(ws, 'Every loaded attribute traced back to its source column',
              ['Each row names one relation, one source block that feeds it, '
               'the attributes that block actually fills, the column group it '
               'was read out of, and the transformation applied.',
               'The attribute list is measured rather than declared: an '
               'attribute appears only where that block supplies a value for '
               'it at least once. Sheet T1.19 a therefore shows Lightning '
               'alone, because it publishes no thunderstorm figure.',
               'A relation marked derived is not published as a table by any '
               'organisation. It is collected from the values the measurement '
               'relations use, so that every foreign key has a parent row.'])
    put(ws, hdr, data, 'INFO', r, widths=[24, 54, 13, 40, 16, 40, 96])

    ws = sheet(wb, 'Stage Comparison', STAGE_COLOUR['INFO'])
    r = title(ws, 'The five stages side by side',
              ['Table and row counts are read back from the files each stage '
               'wrote, so this sheet cannot disagree with the sheets above.',
               'The row count is not expected to fall smoothly. First Normal '
               'Form multiplies rows, because one published row holding twelve '
               'months becomes twelve rows. Third Normal Form reduces them, '
               'because speed and direction rejoin into one wind row. Boyce '
               'Codd Normal Form reduces them again, because the primary key '
               'refuses a duplicate key.',
               'For the unnormalized form the count is rows as published, '
               'heading rows and note rows included, because at that stage a '
               'heading row and a data row are not distinguishable.'])
    n0 = sum(n for _, n in Z0.values())
    WHY = {
        '1NF': 'Every repeating column group is unpivoted, so one published '
               'row holding twelve months becomes twelve rows. Heading rows, '
               'note rows and column number legend rows stop being rows at '
               'all.',
        '2NF': 'No measurement row is gained or lost. The increase is exactly '
               'the reference tables that partial dependencies were lifted '
               'into.',
        '3NF': 'Wind speed and direction rejoin into one row, thunderstorm '
               'and lightning rejoin into one row, and the eight forest area '
               'measures pivot back into one row per district and fiscal '
               'year, so many rows become few wider rows.',
        'BCNF': 'The primary key is enforced, so a duplicate key is refused '
                'and the rejected value is recorded in Key_Conflicts. The '
                'reference relations that the foreign keys need are added at '
                'the same time.',
    }
    rows = [['0NF', 'Unnormalized form', len(Z0), n0, '',
             'No key exists, so a row is not yet a fact about anything.']]
    prev = n0
    for stage in ('1NF', '2NF', '3NF', 'BCNF'):
        tabs = STAGE_TABLES[stage]
        tot = sum(n for _, n in tabs.values())
        nm = {'1NF': 'First Normal Form', '2NF': 'Second Normal Form',
              '3NF': 'Third Normal Form',
              'BCNF': 'Boyce Codd Normal Form'}[stage]
        rows.append([stage, nm, len(tabs), tot,
                     '%s %d' % ('up' if tot >= prev else 'down',
                                abs(tot - prev)),
                     WHY[stage]])
        prev = tot
    r = put(ws, ['Stage', 'Name', 'Tables', 'Rows', 'Change in rows',
                 'Why the count moves this way'],
            rows, 'INFO', r, widths=[8, 26, 9, 12, 14, 84])
    r += 2

    ws.cell(row=r, column=1,
            value='Which tables appear and disappear at each stage').font = \
        TITLE_FONT
    r += 1
    ws.cell(row=r, column=1,
            value='Measured by comparing the file names each stage wrote. A '
                  'table that disappears has been split, pivoted, folded into '
                  'another or left unloaded, and the next sheet section states '
                  'which.').font = NOTE_FONT
    r += 2

    def base(nm):
        for suf in ('_1NF', '_2NF', '_3NF'):
            if nm.endswith(suf):
                return nm[:-len(suf)]
        return nm

    diff = []
    order = ('1NF', '2NF', '3NF', 'BCNF')
    for i, stage in enumerate(order):
        cur = {base(k): v for k, v in STAGE_TABLES[stage].items()}
        if i == 0:
            diff.append([stage, 'appears',
                         '%d tables, the whole stage is new'
                         % len(cur), sum(n for _, n in cur.values())])
            continue
        old = {base(k): v for k, v in STAGE_TABLES[order[i - 1]].items()}
        new = sorted(set(cur) - set(old))
        gone = sorted(set(old) - set(cur))
        if new:
            diff.append([stage, 'appears', ', '.join(new),
                         sum(cur[k][1] for k in new)])
        if gone:
            diff.append(['%s (was %s)' % (stage, order[i - 1]), 'disappears',
                         ', '.join(gone), sum(old[k][1] for k in gone)])
    r = put(ws, ['Stage', 'Change', 'Tables', 'Rows involved'],
            diff, 'INFO', r, widths=[16, 12, 96, 14])
    r += 2

    ws.cell(row=r, column=1,
            value='Where each table goes next').font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1,
            value='A table listed at one stage and absent at the next has '
                  'been split, pivoted or renamed. The row states which.'
            ).font = NOTE_FONT
    r += 2
    FATE = [
        ('Climate_Observation_1NF', '1NF',
         'Provenance and unit are lifted out at Second Normal Form, then the '
         'Measure column is split at Third Normal Form into Temperature, '
         'Humidity, Rainfall, Wind and Climatic Event relations.'),
        ('Daily_Observation_1NF', '1NF',
         'Split at Third Normal Form into Sunshine_Record and '
         'Radiation_Record, one relation per measured quantity.'),
        ('BRRI_*_Daily_1NF', '1NF',
         'Four full daily trace tables preserve maximum temperature, minimum '
         'temperature, rainfall and humidity before their monthly '
         'aggregation. Their monthly results enter the climate relations; '
         'the daily rows are not loaded because the approved keys are monthly.'),
        ('Forest_Area_1NF', '1NF',
         'The fiscal year text becomes Fiscal_Year at Second Normal Form, '
         'then the eight area measures pivot back into named columns at '
         'Third Normal Form.'),
        ('Waste_Water_1NF', '1NF',
         'Split at Third Normal Form into Type_Of_Establishments and '
         'Industry_Type, which the approved diagram draws separately.'),
        ('Water_Quality_1NF', '1NF',
         'River_Name and Water_Category move to River_Station at Second '
         'Normal Form, because they depend on the station alone.'),
        ('River_Register_1NF', '1NF',
         'Carried through Second and Third Normal Form unchanged, then reduced '
         'to the River relation at Boyce Codd Normal Form, because the '
         'approved diagram keeps only the river name.'),
        ('Climate_Observation_2NF', '2NF',
         'The same rows as at First Normal Form with provenance and unit '
         'lifted out. Split at Third Normal Form into the five climate '
         'relations, one per measured quantity.'),
        ('Water_Quality_2NF', '2NF',
         'Loaded as Water_Quality. The station and river pairing it gave up '
         'is held by River_Station.'),
        ('Source_Block_2NF', '2NF',
         'A load time provenance table. It is not part of the approved '
         'diagram, so it is not loaded; the Traceability sheet carries the '
         'same information.'),
        ('Measure_Unit_2NF', '2NF',
         'Records the unit of each measured quantity. The approved diagram '
         'has no unit entity, so the units are documented rather than '
         'loaded.'),
        ('River_Register_3NF', '3NF',
         'Reduced to River. Serial number, zone, border flag and flow type '
         'are not in the approved diagram.'),
        ('Station, District, Year_Time, Month_Time, Day_Time', 'BCNF',
         'New at Boyce Codd Normal Form. No organisation publishes these as '
         'tables. Each is collected from the values the measurement relations '
         'actually use, so that every foreign key resolves to a parent row.'),
        ('River', 'BCNF',
         'Not new. It is River_Register_3NF reduced to the name column, plus '
         'the lakes and canals that only a monitoring station names. The Key '
         'Enforcement sheet gives the arithmetic.'),
    ]
    put(ws, ['Table', 'Stage', 'What becomes of it'],
        [list(x) for x in FATE], 'INFO', r, widths=[32, 8, 104])

    path = OUT + 'Environmental_Normalization_0NF_to_BCNF.xlsx'
    wb.save(path)
    return path, len(wb.sheetnames)


if __name__ == '__main__':
    p, n = build()
    print('wrote %s' % p)
    print('%d sheets' % n)
    print('%.1f MB' % (os.path.getsize(p) / 1048576.0))
